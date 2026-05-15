#!/usr/bin/env python3
"""
LogicMonitor API - Unmanage Interfaces from CSV/XLSX
----------------------------------------------------
Reads a LogicMonitor-style export file and updates interface instance
monitoring/alerting settings for every data row.

This version uses the device/resource ID as the authoritative device selector.
The Resource/displayName column is used only for debug validation/fallback.

Important update behavior:
- LogicMonitor's documented datasource instance update endpoint is PUT.
- The script first GETs the existing instance so required fields are preserved.
- The script then PUTs the existing required fields plus the desired per-row
  values from the CSV:
    "stopMonitoring": true/false
    "disableAlerting": true/false
- If the CSV does not include stopMonitoring or disableAlerting, both default
  to true to preserve the previous unmanage behavior.
- After applying, the script verifies the instance state with a GET.

Sample CSV header:
  Id,Resource,Datasource,Instance,stopMonitoring,disableAlerting

Sample CSV row:
  1,Lenny,Network Interfaces (Linux_SSH_NetworkInterfaces),veth3b28b4af,true,true

Requirements:
- Python 3.8+
- requests
- python-dotenv
- openpyxl (only required for XLSX input)

.env:
ACCESS_ID=your_access_id
ACCESS_KEY=your_access_key
COMPANY=your_company_name

Examples:
  python unmanage_interfaces_from_csv_id_v9.py --input TEST_NEt_Interfaces.csv --debug
  python unmanage_interfaces_from_csv_id_v9.py --input TEST_NEt_Interfaces.csv --apply --debug
"""

import argparse
import base64
import csv
import hashlib
import hmac
import json
import os
import random
import re
import shlex
import time
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import requests
from dotenv import load_dotenv


load_dotenv()

ACCESS_ID = os.getenv("ACCESS_ID", "")
ACCESS_KEY = os.getenv("ACCESS_KEY", "")
COMPANY = os.getenv("COMPANY", "")

BASE_URL = ""
DEBUG_API = False

RETRY_STATUS_CODES = {406, 408, 500, 502, 503, 504}


# -----------------------------
# LogicMonitor auth / requests
# -----------------------------
def validate_creds() -> None:
    global BASE_URL

    if not ACCESS_ID or not ACCESS_KEY or not COMPANY:
        raise SystemExit(
            "Missing ACCESS_ID / ACCESS_KEY / COMPANY. "
            "Set them in your environment or .env file."
        )

    BASE_URL = f"https://{COMPANY}.logicmonitor.com/santaba/rest"


def generate_auth_headers(http_verb: str, resource_path: str, data: str = "") -> Dict[str, str]:
    """
    Generate LogicMonitor LMv1 auth headers.

    The signed data string must exactly match the request body string.
    Query parameters are not included in the signed resourcePath.
    """
    epoch = str(int(time.time() * 1000))
    request_vars = http_verb.upper() + epoch + data + resource_path
    digest = hmac.new(
        ACCESS_KEY.encode("utf-8"),
        msg=request_vars.encode("utf-8"),
        digestmod=hashlib.sha256,
    ).hexdigest()
    signature = base64.b64encode(digest.encode("utf-8")).decode("utf-8")

    return {
        "Authorization": f"LMv1 {ACCESS_ID}:{signature}:{epoch}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }


def build_debug_curl(http_verb: str, full_url: str, data: str = "") -> str:
    parts = [
        "curl",
        "-X",
        http_verb.upper(),
        full_url,
        "-H",
        "Accept: application/json",
        "-H",
        "Content-Type: application/json",
        "-H",
        "Authorization: LMv1 <redacted>",
    ]
    if data:
        parts.extend(["--data", data])

    return " ".join(shlex.quote(str(p)) for p in parts)


def debug_request(http_verb: str, url: str, params: Dict[str, Any], data: str, attempt: int, retries: int) -> None:
    if not DEBUG_API:
        return

    prepared = requests.Request(http_verb.upper(), url, params=params).prepare()
    full_url = prepared.url or url

    print(f"DEBUG API attempt {attempt}/{retries}: {http_verb.upper()} {full_url}")
    if data:
        print(f"DEBUG API payload: {data}")
    print(f"DEBUG API curl: {build_debug_curl(http_verb, full_url, data)}")


def api_raw(
    http_verb: str,
    resource_path: str,
    params: Optional[Dict[str, Any]] = None,
    payload: Optional[Dict[str, Any]] = None,
    retries: int = 3,
    backoff_base: float = 1.0,
) -> str:
    http_verb = http_verb.upper()
    params = params or {}
    url = BASE_URL + resource_path
    data = "" if payload is None else json.dumps(payload, separators=(",", ":"), ensure_ascii=False)

    for attempt in range(1, retries + 1):
        try:
            headers = generate_auth_headers(http_verb, resource_path, data)
            debug_request(http_verb, url, params, data, attempt, retries)

            response = requests.request(
                http_verb,
                url,
                headers=headers,
                params=params,
                data=data if payload is not None else None,
                timeout=60,
            )

            if 200 <= response.status_code < 300:
                return response.text

            body = response.text.strip() or "<empty body>"

            if response.status_code in (401, 403):
                raise RuntimeError(f"{http_verb} {resource_path} failed: {response.status_code} - {body}")

            if response.status_code == 429 and attempt < retries:
                retry_after = response.headers.get("Retry-After")
                sleep_s = 30
                if retry_after:
                    try:
                        sleep_s = int(float(retry_after))
                    except ValueError:
                        pass
                print(
                    f"Rate limited: {http_verb} {resource_path} returned 429 "
                    f"(attempt {attempt}/{retries}). Sleeping {sleep_s}s. Body: {body}"
                )
                time.sleep(sleep_s)
                continue

            if response.status_code in RETRY_STATUS_CODES and attempt < retries:
                sleep_s = backoff_base * (2 ** (attempt - 1)) + random.uniform(0, 0.25)
                print(
                    f"Warning: {http_verb} {resource_path} returned {response.status_code} "
                    f"(attempt {attempt}/{retries}). Sleeping {sleep_s:.2f}s. Body: {body}"
                )
                time.sleep(sleep_s)
                continue

            raise RuntimeError(f"{http_verb} {resource_path} failed: {response.status_code} - {body}")

        except (
            requests.exceptions.Timeout,
            requests.exceptions.ConnectionError,
            requests.exceptions.ChunkedEncodingError,
        ) as exc:
            if attempt < retries:
                sleep_s = backoff_base * (2 ** (attempt - 1)) + random.uniform(0, 0.25)
                print(
                    f"Warning: network error on {http_verb} {resource_path} "
                    f"(attempt {attempt}/{retries}). Sleeping {sleep_s:.2f}s. Error: {exc}"
                )
                time.sleep(sleep_s)
                continue
            raise RuntimeError(f"{http_verb} {resource_path} failed after {retries} attempts: {exc}") from exc

    return ""


def api_json(
    http_verb: str,
    resource_path: str,
    params: Optional[Dict[str, Any]] = None,
    payload: Optional[Dict[str, Any]] = None,
    retries: int = 3,
) -> Dict[str, Any]:
    raw = api_raw(http_verb, resource_path, params=params, payload=payload, retries=retries)
    if not raw.strip():
        return {}

    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"{http_verb.upper()} {resource_path} returned non-JSON content: {exc}") from exc


# -----------------------------
# Generic helpers
# -----------------------------
def normalize_key(value: Any) -> str:
    return str(value or "").strip().lower()


def normalize_col_name(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", normalize_key(value))


def optional_int(value: Any) -> Optional[int]:
    if value is None or str(value).strip() == "":
        return None

    try:
        return int(str(value).strip())
    except ValueError:
        return None


TRUE_VALUES = {"true", "t", "yes", "y", "1", "managed", "manage", "enabled", "enable", "on"}
FALSE_VALUES = {"false", "f", "no", "n", "0", "unmanaged", "unmanage", "disabled", "disable", "off"}


def parse_bool_option(value: Any, default: bool, column_name: str) -> bool:
    """
    Parse true/false style CSV values.

    Empty or missing values return the supplied default.
    """
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    text = str(value).strip().lower()
    if text == "":
        return default

    if text in TRUE_VALUES:
        return True

    if text in FALSE_VALUES:
        return False

    raise ValueError(
        f"Invalid boolean value {value!r} for column {column_name!r}. "
        "Expected true/false, yes/no, 1/0, enabled/disabled, or on/off."
    )


def extract_items(payload: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Support both wrapped LM responses:
      {"data": {"items": [...], "total": 123}}
    and direct responses:
      {"items": [...], "total": 123}
    """
    if not isinstance(payload, dict):
        return [], {}

    data = payload.get("data")
    if isinstance(data, dict) and isinstance(data.get("items"), list):
        return [x for x in data["items"] if isinstance(x, dict)], {
            "total": data.get("total"),
            "searchId": data.get("searchId"),
        }

    items = payload.get("items")
    if isinstance(items, list):
        return [x for x in items if isinstance(x, dict)], {
            "total": payload.get("total"),
            "searchId": payload.get("searchId"),
        }

    return [], {}


def extract_object(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {}

    data = payload.get("data")
    if isinstance(data, dict):
        return data

    return payload


def get_ci(row: Dict[str, Any], *names: str) -> Any:
    normalized = {normalize_col_name(k): v for k, v in row.items()}
    for name in names:
        key = normalize_col_name(name)
        if key in normalized:
            return normalized[key]
    return None


def get_ci_with_source(row: Dict[str, Any], *names: str) -> Tuple[Any, str]:
    normalized = {normalize_col_name(k): (k, v) for k, v in row.items()}
    for name in names:
        key = normalize_col_name(name)
        if key in normalized:
            source, value = normalized[key]
            return value, str(source)
    return None, ""


def escape_lm_filter_value(value: str) -> str:
    return value.replace("\\", "\\\\").replace('"', '\\"')


def build_contains_filter(field: str, value: str) -> str:
    return f'{field}~"{escape_lm_filter_value(value)}"'


def paginate(
    resource_path: str,
    params: Optional[Dict[str, Any]] = None,
    size: int = 200,
    retries: int = 3,
) -> List[Dict[str, Any]]:
    items_all: List[Dict[str, Any]] = []
    offset = 0

    while True:
        page_params = {"size": size, "offset": offset}
        if params:
            page_params.update(params)

        payload = api_json("GET", resource_path, params=page_params, retries=retries)
        items, meta = extract_items(payload)
        items_all.extend(items)

        total = meta.get("total")
        if not items:
            break
        if isinstance(total, int) and total > 0 and len(items_all) >= total:
            break
        if len(items) < size:
            break

        offset += size

    return items_all


def parse_datasource_label(label: str) -> Tuple[str, Optional[str]]:
    """
    Convert:
      Network Interfaces (Linux_SSH_NetworkInterfaces)
    into:
      ("Network Interfaces", "Linux_SSH_NetworkInterfaces")
    """
    label = str(label or "").strip()
    match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", label)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    return label, None


def exactly_one(items: Sequence[Dict[str, Any]], description: str) -> Dict[str, Any]:
    if not items:
        raise RuntimeError(f"No match found for {description}.")

    if len(items) > 1:
        details = "; ".join(
            f"id={x.get('id')} displayName={x.get('displayName') or x.get('dataSourceDisplayName')!r} "
            f"name={x.get('name') or x.get('dataSourceName')!r}"
            for x in items[:10]
        )
        raise RuntimeError(f"Multiple matches found for {description}: {details}")

    return items[0]


# -----------------------------
# Input loading
# -----------------------------
def file_is_xlsx(path: Path) -> bool:
    with path.open("rb") as fh:
        return fh.read(4) == b"PK\x03\x04"


def read_xlsx_rows(path: Path) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("Input appears to be XLSX. Install openpyxl: pip install openpyxl") from exc

    workbook = load_workbook(BytesIO(path.read_bytes()), read_only=True, data_only=True)
    sheet = workbook.active
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def read_csv_rows(path: Path) -> List[List[Any]]:
    text = path.read_bytes().decode("utf-8-sig", errors="replace")
    sample = text[:4096]

    try:
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    try:
        return [row for row in csv.reader(text.splitlines(), dialect=dialect)]
    except (csv.Error, ValueError):
        return [row for row in csv.reader(text.splitlines(), dialect=csv.excel)]


def rows_to_dicts(rows: List[List[Any]], required_columns: Sequence[str]) -> Tuple[List[Dict[str, Any]], int]:
    required = {normalize_col_name(c) for c in required_columns}

    header_idx: Optional[int] = None
    header: List[str] = []

    for idx, row in enumerate(rows):
        found = {normalize_col_name(cell) for cell in row if cell is not None and str(cell).strip()}
        if required.issubset(found):
            header_idx = idx
            header = [str(cell).strip() if cell is not None else "" for cell in row]
            break

    if header_idx is None:
        raise RuntimeError(
            "Could not find a header row containing required columns: "
            + ", ".join(required_columns)
        )

    records: List[Dict[str, Any]] = []
    for raw_row in rows[header_idx + 1 :]:
        if not any(cell is not None and str(cell).strip() for cell in raw_row):
            continue

        padded = list(raw_row) + [None] * max(0, len(header) - len(raw_row))
        records.append({header[i]: padded[i] for i in range(len(header))})

    return records, header_idx + 1


def load_input_records(path: Path, args: argparse.Namespace) -> Tuple[List[Dict[str, Any]], int]:
    rows = read_xlsx_rows(path) if file_is_xlsx(path) else read_csv_rows(path)
    required = [args.device_column, args.datasource_column, args.instance_column]
    return rows_to_dicts(rows, required)


# -----------------------------
# LogicMonitor resolution
# -----------------------------
class Resolver:
    def __init__(self, size: int, retries: int):
        self.size = size
        self.retries = retries
        self.device_cache: Dict[str, Dict[str, Any]] = {}
        self.datasource_cache: Dict[Tuple[int, str], Dict[str, Any]] = {}
        self.instance_cache: Dict[Tuple[int, int, str], Dict[str, Any]] = {}

    def get_device_by_id(self, device_id: int) -> Dict[str, Any]:
        cache_key = f"id:{device_id}"
        if cache_key in self.device_cache:
            return self.device_cache[cache_key]

        payload = api_json(
            "GET",
            f"/device/devices/{device_id}",
            params={"fields": "id,displayName,name"},
            retries=self.retries,
        )
        device = extract_object(payload)

        if optional_int(device.get("id")) != device_id:
            raise RuntimeError(f"Device ID {device_id} was not returned by the API.")

        self.device_cache[cache_key] = device
        return device

    def debug_check_device_id(self, device_id: int, expected_name: str = "") -> None:
        if not DEBUG_API:
            return

        print(f"DEBUG device id check: verifying deviceId={device_id}")
        device = self.get_device_by_id(device_id)
        display_name = str(device.get("displayName") or "")
        name = str(device.get("name") or "")

        print(f"DEBUG device id check: found id={device.get('id')} displayName={display_name!r} name={name!r}")

        if expected_name:
            expected = normalize_key(expected_name)
            if expected in {normalize_key(display_name), normalize_key(name)}:
                print(f"DEBUG device name check: Resource value matches deviceId={device_id}")
            else:
                print(
                    "DEBUG device name check WARNING: Resource value "
                    f"{expected_name!r} did not exactly match displayName {display_name!r} or name {name!r}; "
                    f"continuing with deviceId={device_id}"
                )

    def resolve_device_by_name_for_debug(self, device_name_or_display_name: str) -> Dict[str, Any]:
        requested = str(device_name_or_display_name or "").strip()
        if not requested:
            raise RuntimeError("Device name/displayName is empty.")

        key = normalize_key(requested)
        cache_key = f"name:{key}"
        if cache_key in self.device_cache:
            return self.device_cache[cache_key]

        fields = "id,displayName,name"
        candidates_by_id: Dict[int, Dict[str, Any]] = {}

        for field in ("displayName", "name"):
            for item in paginate(
                "/device/devices",
                params={"fields": fields, "filter": build_contains_filter(field, requested)},
                size=self.size,
                retries=self.retries,
            ):
                item_id = optional_int(item.get("id"))
                if item_id is not None:
                    candidates_by_id[item_id] = item

        candidates = list(candidates_by_id.values())

        if DEBUG_API:
            print(f"DEBUG device name fallback: requested displayName/name={requested!r}")
            for item in candidates[:20]:
                print(
                    "DEBUG device name fallback candidate: "
                    f"id={item.get('id')} displayName={item.get('displayName')!r} name={item.get('name')!r}"
                )
            if not candidates:
                print("DEBUG device name fallback: API returned 0 candidates")

        exact = [
            item
            for item in candidates
            if normalize_key(item.get("displayName")) == key or normalize_key(item.get("name")) == key
        ]

        device = exactly_one(exact, f"device displayName/name {requested!r}")
        self.device_cache[cache_key] = device
        self.device_cache[f"id:{device.get('id')}"] = device
        return device

    def resolve_device_datasource(self, device_id: int, datasource_label: str) -> Dict[str, Any]:
        ds_display, ds_name = parse_datasource_label(datasource_label)
        cache_key = (device_id, normalize_key(datasource_label))
        if cache_key in self.datasource_cache:
            return self.datasource_cache[cache_key]

        fields = "id,dataSourceId,dataSourceName,dataSourceDisplayName,stopMonitoring"

        ds_display_norm = normalize_key(ds_display)
        ds_name_norm = normalize_key(ds_name)
        full_label_norm = normalize_key(datasource_label)

        ds_display_loose = normalize_col_name(ds_display)
        ds_name_loose = normalize_col_name(ds_name)
        full_label_loose = normalize_col_name(datasource_label)

        def datasource_rank(item: Dict[str, Any]) -> Optional[Tuple[int, str]]:
            item_display_raw = item.get("dataSourceDisplayName")
            item_name_raw = item.get("dataSourceName")

            item_display = normalize_key(item_display_raw)
            item_name = normalize_key(item_name_raw)
            item_full = normalize_key(f"{item_display_raw or ''} ({item_name_raw or ''})")

            item_display_loose = normalize_col_name(item_display_raw)
            item_name_loose = normalize_col_name(item_name_raw)
            item_full_loose = normalize_col_name(f"{item_display_raw or ''} ({item_name_raw or ''})")

            if ds_name_norm:
                if item_name == ds_name_norm:
                    return (0, "exact dataSourceName/module-name match")
                if ds_name_loose and item_name_loose == ds_name_loose:
                    return (1, "loose dataSourceName/module-name match")
                if item_full == full_label_norm:
                    return (2, "exact full datasource label match")
                if item_full_loose == full_label_loose:
                    return (3, "loose full datasource label match")
                if item_display == ds_name_norm:
                    return (4, "displayName matched requested module name")
                if ds_name_loose and item_display_loose == ds_name_loose:
                    return (5, "loose displayName matched requested module name")

                # Last resort only. This is commonly ambiguous for Network Interfaces.
                if item_display == ds_display_norm:
                    return (50, "displayName-only fallback match")
                if ds_display_loose and item_display_loose == ds_display_loose:
                    return (51, "loose displayName-only fallback match")
                return None

            if item_display == ds_display_norm:
                return (0, "exact displayName match")
            if ds_display_loose and item_display_loose == ds_display_loose:
                return (1, "loose displayName match")
            if item_name == ds_display_norm:
                return (2, "exact dataSourceName match")
            if ds_display_loose and item_name_loose == ds_display_loose:
                return (3, "loose dataSourceName match")

            return None

        def select_best(candidates: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
            ranked: List[Tuple[int, str, Dict[str, Any]]] = []
            seen_ids = set()

            for item in candidates:
                item_id = optional_int(item.get("id"))
                if item_id is not None:
                    if item_id in seen_ids:
                        continue
                    seen_ids.add(item_id)

                rank = datasource_rank(item)
                if rank:
                    ranked.append((rank[0], rank[1], item))

            if DEBUG_API and ranked:
                print("DEBUG datasource ranked candidate(s):")
                for rank, reason, item in sorted(ranked, key=lambda r: (r[0], optional_int(r[2].get("id")) or 0))[:25]:
                    print(
                        "DEBUG datasource candidate: "
                        f"rank={rank} reason={reason!r} hdsId={item.get('id')} "
                        f"dataSourceId={item.get('dataSourceId')} "
                        f"displayName={item.get('dataSourceDisplayName')!r} "
                        f"name={item.get('dataSourceName')!r} "
                        f"stopMonitoring={item.get('stopMonitoring')!r}"
                    )

            if not ranked:
                raise RuntimeError(f"No match found for datasource {datasource_label!r} on device id={device_id}.")

            ranked.sort(key=lambda r: (r[0], optional_int(r[2].get("id")) or 0))
            best_rank = ranked[0][0]
            best = [entry for entry in ranked if entry[0] == best_rank]

            if len(best) == 1:
                selected = best[0][2]
                if DEBUG_API:
                    print(
                        "DEBUG datasource selected: "
                        f"hdsId={selected.get('id')} dataSourceId={selected.get('dataSourceId')} "
                        f"displayName={selected.get('dataSourceDisplayName')!r} "
                        f"name={selected.get('dataSourceName')!r} "
                        f"rank={best[0][0]} reason={best[0][1]!r}"
                    )
                return selected

            details = "; ".join(
                f"hdsId={entry[2].get('id')} dataSourceId={entry[2].get('dataSourceId')} "
                f"displayName={entry[2].get('dataSourceDisplayName')!r} "
                f"name={entry[2].get('dataSourceName')!r} reason={entry[1]!r}"
                for entry in best[:20]
            )
            raise RuntimeError(
                f"Multiple equally ranked datasource matches for {datasource_label!r} on device id={device_id}: "
                f"{details}. Add hdsId/deviceDataSourceId to the CSV to target the datasource directly."
            )

        filtered_candidates: Dict[int, Dict[str, Any]] = {}
        lookup_order: List[Tuple[str, str]] = []

        if ds_name:
            lookup_order.extend([("dataSourceName", ds_name), ("dataSourceDisplayName", ds_name)])
        if ds_display:
            lookup_order.extend([("dataSourceDisplayName", ds_display), ("dataSourceName", ds_display)])

        for field, value in lookup_order:
            items = paginate(
                f"/device/devices/{device_id}/devicedatasources",
                params={"fields": fields, "filter": build_contains_filter(field, value)},
                size=self.size,
                retries=self.retries,
            )
            if DEBUG_API:
                print(f"DEBUG datasource filtered lookup: field={field} value={value!r} returned {len(items)} item(s)")

            for item in items:
                item_id = optional_int(item.get("id"))
                if item_id is not None:
                    filtered_candidates[item_id] = item

        try:
            selected = select_best(list(filtered_candidates.values()))
            self.datasource_cache[cache_key] = selected
            return selected
        except RuntimeError as exc:
            if DEBUG_API:
                print(f"DEBUG datasource filtered lookup did not select a datasource: {exc}")

        all_items = paginate(
            f"/device/devices/{device_id}/devicedatasources",
            params={"fields": fields},
            size=self.size,
            retries=self.retries,
        )

        if DEBUG_API:
            print(f"DEBUG datasource unfiltered fallback: deviceId={device_id} has {len(all_items)} assigned datasource(s)")
            for item in all_items[:50]:
                print(
                    "DEBUG datasource available: "
                    f"hdsId={item.get('id')} dataSourceId={item.get('dataSourceId')} "
                    f"displayName={item.get('dataSourceDisplayName')!r} "
                    f"name={item.get('dataSourceName')!r} "
                    f"stopMonitoring={item.get('stopMonitoring')!r}"
                )
            if len(all_items) > 50:
                print(f"DEBUG datasource unfiltered fallback: {len(all_items) - 50} additional datasource(s) not shown")

        selected = select_best(all_items)
        self.datasource_cache[cache_key] = selected
        return selected

    def get_instance_by_id(
        self,
        device_id: int,
        hds_id: int,
        instance_id: int,
        refresh: bool = False,
    ) -> Dict[str, Any]:
        cache_key = (device_id, hds_id, f"id:{instance_id}")
        if not refresh and cache_key in self.instance_cache:
            return self.instance_cache[cache_key]

        fields = "id,name,displayName,wildValue,wildValue2,groupId,description,stopMonitoring,disableAlerting"
        payload = api_json(
            "GET",
            f"/device/devices/{device_id}/devicedatasources/{hds_id}/instances/{instance_id}",
            params={"fields": fields},
            retries=self.retries,
        )
        instance = extract_object(payload)

        # If the fields projection does not return groupId in a given portal/API
        # version, retry without fields so the PUT payload can preserve it.
        if optional_int(instance.get("id")) == instance_id and optional_int(instance.get("groupId")) is None:
            payload = api_json(
                "GET",
                f"/device/devices/{device_id}/devicedatasources/{hds_id}/instances/{instance_id}",
                retries=self.retries,
            )
            instance = extract_object(payload)

        if optional_int(instance.get("id")) != instance_id:
            raise RuntimeError(
                f"Instance ID {instance_id} was not returned by the API for deviceId={device_id}, hdsId={hds_id}."
            )

        self.instance_cache[cache_key] = instance
        return instance

    def resolve_instance(self, device_id: int, hds_id: int, instance_name: str) -> Dict[str, Any]:
        if not str(instance_name or "").strip():
            raise RuntimeError("Instance name is empty.")

        cache_key = (device_id, hds_id, normalize_key(instance_name))
        if cache_key in self.instance_cache:
            return self.instance_cache[cache_key]

        fields = "id,name,displayName,wildValue,wildValue2,groupId,description,stopMonitoring,disableAlerting"
        resource_path = f"/device/devices/{device_id}/devicedatasources/{hds_id}/instances"

        target = normalize_key(instance_name)
        target_loose = normalize_col_name(instance_name)

        def instance_rank(item: Dict[str, Any]) -> Optional[Tuple[int, str]]:
            values = {
                "displayName": item.get("displayName"),
                "wildValue": item.get("wildValue"),
                "wildValue2": item.get("wildValue2"),
                "name": item.get("name"),
            }
            normalized = {field: normalize_key(value) for field, value in values.items()}
            loose = {field: normalize_col_name(value) for field, value in values.items()}

            exact_order = [
                ("displayName", "exact displayName match"),
                ("wildValue", "exact wildValue match"),
                ("wildValue2", "exact wildValue2 match"),
                ("name", "exact name match"),
            ]
            for rank, (field, reason) in enumerate(exact_order):
                if normalized[field] == target:
                    return (rank, reason)

            loose_order = [
                ("displayName", "loose displayName match"),
                ("wildValue", "loose wildValue match"),
                ("wildValue2", "loose wildValue2 match"),
                ("name", "loose name match"),
            ]
            for offset, (field, reason) in enumerate(loose_order):
                if target_loose and loose[field] == target_loose:
                    return (10 + offset, reason)

            contains_order = [
                ("displayName", "displayName contains requested instance"),
                ("wildValue", "wildValue contains requested instance"),
                ("wildValue2", "wildValue2 contains requested instance"),
                ("name", "name contains requested instance"),
            ]
            for offset, (field, reason) in enumerate(contains_order):
                if target and target in normalized[field]:
                    return (50 + offset, reason)
                if target_loose and target_loose in loose[field]:
                    return (60 + offset, f"loose {reason}")

            return None

        def format_instance(item: Dict[str, Any]) -> str:
            return (
                f"instanceId={item.get('id')} "
                f"displayName={item.get('displayName')!r} "
                f"name={item.get('name')!r} "
                f"wildValue={item.get('wildValue')!r} "
                f"wildValue2={item.get('wildValue2')!r} "
                f"groupId={item.get('groupId')!r} "
                f"stopMonitoring={item.get('stopMonitoring')!r} "
                f"disableAlerting={item.get('disableAlerting')!r}"
            )

        def select_best(candidates: Sequence[Dict[str, Any]]) -> Dict[str, Any]:
            ranked: List[Tuple[int, str, Dict[str, Any]]] = []
            seen_ids = set()

            for item in candidates:
                item_id = optional_int(item.get("id"))
                if item_id is not None:
                    if item_id in seen_ids:
                        continue
                    seen_ids.add(item_id)

                rank = instance_rank(item)
                if rank:
                    ranked.append((rank[0], rank[1], item))

            if DEBUG_API and ranked:
                print("DEBUG instance ranked candidate(s):")
                for rank, reason, item in sorted(ranked, key=lambda r: (r[0], optional_int(r[2].get("id")) or 0))[:25]:
                    print(f"DEBUG instance candidate: rank={rank} reason={reason!r} {format_instance(item)}")

            if not ranked:
                raise RuntimeError(
                    f"No match found for instance {instance_name!r} on device id={device_id}, "
                    f"deviceDatasource id={hds_id}."
                )

            ranked.sort(key=lambda r: (r[0], optional_int(r[2].get("id")) or 0))
            best_rank = ranked[0][0]
            best = [entry for entry in ranked if entry[0] == best_rank]

            if len(best) == 1:
                selected = best[0][2]
                if DEBUG_API:
                    print(f"DEBUG instance selected: {format_instance(selected)} rank={best[0][0]} reason={best[0][1]!r}")
                return selected

            details = "; ".join(format_instance(entry[2]) + f" reason={entry[1]!r}" for entry in best[:20])
            raise RuntimeError(
                f"Multiple equally ranked instance matches for {instance_name!r} on device id={device_id}, "
                f"deviceDatasource id={hds_id}: {details}. Add instanceId/dataSourceInstanceId to the CSV."
            )

        filtered_candidates: Dict[int, Dict[str, Any]] = {}

        for field in ("displayName", "wildValue", "wildValue2", "name"):
            items = paginate(
                resource_path,
                params={"fields": fields, "filter": build_contains_filter(field, instance_name)},
                size=self.size,
                retries=self.retries,
            )
            if DEBUG_API:
                print(f"DEBUG instance filtered lookup: field={field} value={instance_name!r} returned {len(items)} item(s)")

            for item in items:
                item_id = optional_int(item.get("id"))
                if item_id is not None:
                    filtered_candidates[item_id] = item

        try:
            selected = select_best(list(filtered_candidates.values()))
            selected = self.get_instance_by_id(device_id, hds_id, int(selected["id"]), refresh=True)
            self.instance_cache[cache_key] = selected
            return selected
        except RuntimeError as exc:
            if DEBUG_API:
                print(f"DEBUG instance filtered lookup did not select an instance: {exc}")

        all_items = paginate(
            resource_path,
            params={"fields": fields},
            size=self.size,
            retries=self.retries,
        )

        if DEBUG_API:
            print(f"DEBUG instance unfiltered fallback: deviceId={device_id} hdsId={hds_id} has {len(all_items)} instance(s)")

            related = []
            for item in all_items:
                joined_norm = " ".join(
                    normalize_key(item.get(field))
                    for field in ("displayName", "wildValue", "wildValue2", "name")
                )
                joined_loose = " ".join(
                    normalize_col_name(item.get(field))
                    for field in ("displayName", "wildValue", "wildValue2", "name")
                )
                if (target and target in joined_norm) or (target_loose and target_loose in joined_loose):
                    related.append(item)

            shown = related if related else all_items
            if related:
                print(f"DEBUG instance unfiltered fallback: {len(related)} candidate(s) contain the requested value")
            elif all_items:
                print(f"DEBUG instance unfiltered fallback: no values contained {instance_name!r}; showing first 50")
            else:
                print("DEBUG instance unfiltered fallback: API returned 0 instances")

            for item in shown[:50]:
                print(f"DEBUG instance available: {format_instance(item)}")
            if len(shown) > 50:
                print(f"DEBUG instance unfiltered fallback: {len(shown) - 50} additional instance(s) not shown")

        selected = select_best(all_items)
        selected = self.get_instance_by_id(device_id, hds_id, int(selected["id"]), refresh=True)
        self.instance_cache[cache_key] = selected
        return selected


# -----------------------------
# Instance update
# -----------------------------
def build_instance_update_payload(
    instance: Dict[str, Any],
    desired_stop_monitoring: bool,
    desired_disable_alerting: bool,
) -> Dict[str, Any]:
    """
    Build the full PUT payload for a datasource instance update.

    PUT can reset omitted fields to defaults, so this preserves existing
    required fields and only changes stopMonitoring/disableAlerting to the
    desired per-row CSV values.
    """
    group_id = optional_int(instance.get("groupId"))
    display_name = str(instance.get("displayName") or "").strip()
    wild_value = str(instance.get("wildValue") or "").strip()

    if group_id is None:
        raise RuntimeError(
            "Cannot update instance because groupId was not returned by the API. "
            "The datasource instance PUT endpoint requires groupId."
        )

    if not display_name:
        display_name = str(instance.get("name") or instance.get("id") or "").strip()

    if not wild_value:
        raise RuntimeError(
            "Cannot update instance because wildValue was not returned by the API. "
            "The datasource instance PUT endpoint requires wildValue."
        )

    payload: Dict[str, Any] = {
        "groupId": group_id,
        "displayName": display_name,
        "wildValue": wild_value,
        "stopMonitoring": desired_stop_monitoring,
        "disableAlerting": desired_disable_alerting,
    }

    wild_value_2 = instance.get("wildValue2")
    if wild_value_2 is not None and str(wild_value_2).strip() != "":
        payload["wildValue2"] = wild_value_2

    description = instance.get("description")
    if description is not None and str(description).strip() != "":
        payload["description"] = description

    return payload


def put_instance_monitoring_state(
    device_id: int,
    hds_id: int,
    instance_id: int,
    instance: Dict[str, Any],
    desired_stop_monitoring: bool,
    desired_disable_alerting: bool,
    retries: int,
) -> Dict[str, Any]:
    payload = build_instance_update_payload(
        instance=instance,
        desired_stop_monitoring=desired_stop_monitoring,
        desired_disable_alerting=desired_disable_alerting,
    )
    return api_json(
        "PUT",
        f"/device/devices/{device_id}/devicedatasources/{hds_id}/instances/{instance_id}",
        payload=payload,
        retries=retries,
    )


def verify_instance_monitoring_state(
    resolver: Resolver,
    device_id: int,
    hds_id: int,
    instance_id: int,
    desired_stop_monitoring: bool,
    desired_disable_alerting: bool,
    attempts: int = 3,
    sleep_s: float = 1.0,
) -> Dict[str, Any]:
    latest: Dict[str, Any] = {}

    for attempt in range(1, attempts + 1):
        latest = resolver.get_instance_by_id(device_id, hds_id, instance_id, refresh=True)

        if (
            latest.get("stopMonitoring") is desired_stop_monitoring
            and latest.get("disableAlerting") is desired_disable_alerting
        ):
            if DEBUG_API:
                print(
                    "DEBUG verify instance state: "
                    f"stopMonitoring={latest.get('stopMonitoring')!r} "
                    f"disableAlerting={latest.get('disableAlerting')!r}"
                )
            return latest

        if attempt < attempts:
            time.sleep(sleep_s)

    raise RuntimeError(
        "Update request completed, but verification GET did not show the requested state. "
        f"Requested stopMonitoring={desired_stop_monitoring!r}, "
        f"disableAlerting={desired_disable_alerting!r}. Current values: "
        f"stopMonitoring={latest.get('stopMonitoring')!r}, "
        f"disableAlerting={latest.get('disableAlerting')!r}"
    )


# -----------------------------
# Main processing
# -----------------------------
def process_file(args: argparse.Namespace) -> int:
    global DEBUG_API
    DEBUG_API = bool(args.debug)

    validate_creds()

    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    records, header_row = load_input_records(input_path, args)
    print(f"Loaded {len(records)} data rows from {input_path} (header row: {header_row}).")

    resolver = Resolver(size=args.size, retries=args.retries)

    results: List[Dict[str, Any]] = []
    targeted = 0
    skipped = 0
    failed = 0

    for row_num, row in enumerate(records, start=header_row + 1):
        device_name = str(get_ci(row, args.device_column) or "").strip()
        datasource_label = str(get_ci(row, args.datasource_column) or "").strip()
        instance_name = str(get_ci(row, args.instance_column) or "").strip()

        stop_monitoring_value = get_ci(row, args.stop_monitoring_column)
        disable_alerting_value = get_ci(row, args.disable_alerting_column)

        try:
            desired_stop_monitoring = parse_bool_option(
                stop_monitoring_value,
                default=True,
                column_name=args.stop_monitoring_column,
            )
            desired_disable_alerting = parse_bool_option(
                disable_alerting_value,
                default=True,
                column_name=args.disable_alerting_column,
            )
        except ValueError as exc:
            failed += 1
            results.append(
                {
                    "row": row_num,
                    "resource": device_name,
                    "datasource": datasource_label,
                    "instance": instance_name,
                    "action": "failed",
                    "error": str(exc),
                }
            )
            print(f"ERROR row={row_num}: {exc}")
            continue

        device_id_value, device_id_source = get_ci_with_source(
            row,
            "deviceId",
            "device id",
            "resourceId",
            "resource id",
            args.device_id_column,
        )
        device_id = optional_int(device_id_value)
        hds_id = optional_int(get_ci(row, "hdsId", "deviceDataSourceId", "deviceDatasourceId"))
        instance_id = optional_int(get_ci(row, "instanceId", "dataSourceInstanceId"))

        result: Dict[str, Any] = {
            "row": row_num,
            "resource": device_name,
            "datasource": datasource_label,
            "instance": instance_name,
            "desiredStopMonitoring": desired_stop_monitoring,
            "desiredDisableAlerting": desired_disable_alerting,
            "deviceId": device_id,
            "deviceIdSource": device_id_source if device_id else "",
            "hdsId": hds_id,
            "instanceId": instance_id,
            "action": "",
            "error": "",
        }

        try:
            if device_id:
                resolver.debug_check_device_id(device_id, expected_name=device_name)
            elif args.debug and device_name:
                print(
                    "DEBUG device id check: no device ID found in row; "
                    f"checking {args.device_column}={device_name!r} against device displayName/name"
                )
                device = resolver.resolve_device_by_name_for_debug(device_name)
                device_id = int(device["id"])
                result["deviceId"] = device_id
                result["deviceIdSource"] = "debug_name_lookup"
                print(f"DEBUG device name fallback: found deviceId={device_id}; continuing")
            else:
                raise RuntimeError(
                    f"Missing device ID. Expected a numeric value in {args.device_id_column!r}, "
                    "deviceId, or resourceId. Run with --debug to try exact Resource displayName/name fallback."
                )

            if not hds_id:
                if not datasource_label:
                    raise RuntimeError(
                        f"Missing {args.datasource_column} and no hdsId/deviceDataSourceId value."
                    )
                datasource = resolver.resolve_device_datasource(device_id, datasource_label)
                hds_id = int(datasource["id"])
                result["hdsId"] = hds_id

            if not instance_id:
                if not instance_name:
                    raise RuntimeError(
                        f"Missing {args.instance_column} and no instanceId/dataSourceInstanceId value."
                    )
                instance = resolver.resolve_instance(device_id, hds_id, instance_name)
                instance_id = int(instance["id"])
                result["instanceId"] = instance_id
            else:
                instance = resolver.get_instance_by_id(device_id, hds_id, instance_id, refresh=True)

            current_stop_monitoring = instance.get("stopMonitoring")
            current_disable_alerting = instance.get("disableAlerting")

            if (
                current_stop_monitoring is desired_stop_monitoring
                and current_disable_alerting is desired_disable_alerting
            ):
                skipped += 1
                result["action"] = "already_desired_state"
                results.append(result)
                print(
                    f"Already in desired state: row={row_num} "
                    f"deviceId={device_id} hdsId={hds_id} instanceId={instance_id} "
                    f"stopMonitoring={desired_stop_monitoring} "
                    f"disableAlerting={desired_disable_alerting}"
                )
                continue

            print(
                f"{'Would update' if not args.apply else 'Updating'}: "
                f"row={row_num} deviceId={device_id} hdsId={hds_id} "
                f"instanceId={instance_id} instance={instance_name} "
                f"stopMonitoring {current_stop_monitoring!r}->{desired_stop_monitoring!r} "
                f"disableAlerting {current_disable_alerting!r}->{desired_disable_alerting!r}"
            )

            if args.apply:
                put_instance_monitoring_state(
                    device_id=device_id,
                    hds_id=hds_id,
                    instance_id=instance_id,
                    instance=instance,
                    desired_stop_monitoring=desired_stop_monitoring,
                    desired_disable_alerting=desired_disable_alerting,
                    retries=args.retries,
                )
                verify_instance_monitoring_state(
                    resolver=resolver,
                    device_id=device_id,
                    hds_id=hds_id,
                    instance_id=instance_id,
                    desired_stop_monitoring=desired_stop_monitoring,
                    desired_disable_alerting=desired_disable_alerting,
                )
                result["action"] = "updated"
            else:
                result["action"] = "dry_run"

            targeted += 1

            if args.sleep > 0:
                time.sleep(args.sleep)

        except Exception as exc:
            failed += 1
            result["action"] = "failed"
            result["error"] = str(exc)
            print(f"ERROR row={row_num}: {exc}")

        results.append(result)

    if args.results:
        write_results_csv(Path(args.results), results)
        print(f"Saved results -> {args.results}")

    mode = "APPLY" if args.apply else "DRY-RUN"
    print(f"\nDone. Mode={mode}; targeted={targeted}; skipped={skipped}; failed={failed}")
    return 1 if failed else 0


def write_results_csv(path: Path, results: List[Dict[str, Any]]) -> None:
    if not results:
        return

    fieldnames: List[str] = []
    for result in results:
        for key in result.keys():
            if key not in fieldnames:
                fieldnames.append(key)

    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Update LogicMonitor interface instance stopMonitoring/disableAlerting "
            "values for every CSV row."
        )
    )

    parser.add_argument("--input", required=True, help="Input CSV or XLSX file. XLSX may have a .csv extension.")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Actually PUT updates to instances. Without this flag, the script runs in dry-run mode.",
    )
    parser.add_argument("--results", default="unmanage_interfaces_results.csv", help="Results CSV path.")
    parser.add_argument("--size", type=int, default=200, help="Page size for list endpoints.")
    parser.add_argument("--sleep", type=float, default=0.0, help="Sleep seconds between successful updates.")
    parser.add_argument("--retries", type=int, default=3, help="Retry attempts for API calls.")
    parser.add_argument(
        "--debug",
        action="store_true",
        help="Print API requests, payloads, redacted curl commands, and resolution checks.",
    )

    parser.add_argument(
        "--device-id-column",
        default="Id",
        help="Device/resource ID column name. Defaults to LogicMonitor export column 'Id'.",
    )
    parser.add_argument(
        "--device-column",
        default="Resource",
        help="Device displayName/name column used for debug validation and fallback only.",
    )
    parser.add_argument("--datasource-column", default="Datasource", help="Datasource column name.")
    parser.add_argument("--instance-column", default="Instance", help="Instance/interface column name.")
    parser.add_argument(
        "--stop-monitoring-column",
        default="stopMonitoring",
        help=(
            "CSV column containing desired stopMonitoring true/false value. "
            "Defaults to true when the column is missing or blank."
        ),
    )
    parser.add_argument(
        "--disable-alerting-column",
        default="disableAlerting",
        help=(
            "CSV column containing desired disableAlerting true/false value. "
            "Defaults to true when the column is missing or blank."
        ),
    )

    return parser.parse_args()


def main() -> None:
    args = parse_args()
    raise SystemExit(process_file(args))


if __name__ == "__main__":
    main()
